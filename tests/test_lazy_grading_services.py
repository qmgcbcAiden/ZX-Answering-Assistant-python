"""plugins/one_click_rating_for_projects 的 GradingService / TemplateService / ExcelExporter 单元测试。

这些 service 从 LazyAIGradingView 抽出，本测试解锁了之前无法覆盖的评分算法/批语CRUD/导出逻辑。
"""

import json
import os
import sys
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parents[1]
PLUGINS_DIR = ROOT / "plugins"
if str(PLUGINS_DIR) not in sys.path:
    sys.path.insert(0, str(PLUGINS_DIR))

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from one_click_rating_for_projects.grading_service import GradingService
from one_click_rating_for_projects.template_service import TemplateService
from one_click_rating_for_projects.excel_exporter import ExcelExporter
from one_click_rating_for_projects import scoring


class FakeStudent:
    """模拟 models.ProjectResult（GradingService 访问的字段）。"""

    def __init__(self, sid, name, screenshot_count=5, desc_char_count=200,
                 has_attachment=True, log_stage_count=3, log_total_chars=300):
        self.id = sid
        self.student_name = name
        self.screenshot_count = screenshot_count
        self.desc_char_count = desc_char_count
        self.has_attachment = has_attachment
        self.log_stage_count = log_stage_count
        self.log_total_chars = log_total_chars
        self.pro_score = None
        self.review_comments = None
        self.commit_logs_raw = None


class FakeApiClient:
    """模拟 LazyGradingAPIClient。"""

    def __init__(self, detail=None, fail_ids=None):
        self.detail = detail if detail is not None else {"commitLogs": ["log1"]}
        self.fail_ids = set(fail_ids or [])
        self.audit_calls = []

    def get_student_result_with_logs(self, rid):
        if rid in self.fail_ids:
            raise RuntimeError("mock fetch fail")
        return self.detail

    def audit_result(self, rid, pro_score, review_comments):
        self.audit_calls.append({"rid": rid, "pro_score": pro_score, "review_comments": review_comments})


class _TemplatesIsolatedTest(unittest.TestCase):
    """隔离 scoring._TEMPLATES_FILE 到临时文件，避免读写插件真实文件。"""

    def setUp(self):
        fd, path = tempfile.mkstemp(suffix=".json")
        os.close(fd)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(
                {"short": ["默认短评1", "默认短评2"], "long": ["默认长评1"]},
                f, ensure_ascii=False,
            )
        self._tmp_path = path
        self._orig = scoring._TEMPLATES_FILE
        scoring._TEMPLATES_FILE = Path(path)

    def tearDown(self):
        scoring._TEMPLATES_FILE = self._orig
        try:
            os.remove(self._tmp_path)
        except OSError:
            pass


class GradingServiceTests(_TemplatesIsolatedTest):
    def test_grade_groups_high_success(self):
        s1 = FakeStudent("1", "张三")
        s2 = FakeStudent("2", "李四")
        api = FakeApiClient()
        progress = []
        stats = GradingService(api).grade_groups(
            [("项目A", [s1, s2])],
            strictness="high",
            comment_picker=scoring.CommentPicker(),
            on_progress=progress.append,
        )
        self.assertEqual(stats["graded"], 2)
        self.assertEqual(stats["failed"], 0)
        self.assertEqual(len(api.audit_calls), 2)
        # 分数/批语写回 student
        self.assertIsNotNone(s1.pro_score)
        self.assertIsNotNone(s1.review_comments)
        # commit_logs_raw 写回
        self.assertEqual(s1.commit_logs_raw, ["log1"])
        # 进度文案
        self.assertTrue(any("分析 张三" in p for p in progress))
        self.assertTrue(any("提交" in p for p in progress))

    def test_grade_groups_records_failure(self):
        s1 = FakeStudent("1", "张三")
        s2 = FakeStudent("2", "李四")
        api = FakeApiClient(fail_ids={"1"})  # s1 拉详情失败
        stats = GradingService(api).grade_groups(
            [("项目A", [s1, s2])],
            strictness="high",
            comment_picker=scoring.CommentPicker(),
        )
        self.assertEqual(stats["graded"], 1)  # s2 成功
        self.assertEqual(stats["failed"], 1)  # s1 失败
        self.assertTrue(any("张三" in n for n in stats["failed_names"]))

    def test_grade_groups_skips_empty_group(self):
        api = FakeApiClient()
        stats = GradingService(api).grade_groups(
            [("空组", [])],
            strictness="high",
            comment_picker=scoring.CommentPicker(),
        )
        self.assertEqual(stats["graded"], 0)
        self.assertEqual(len(api.audit_calls), 0)

    def test_grade_groups_accumulates_stats_across_groups(self):
        api = FakeApiClient()
        stats = GradingService(api).grade_groups(
            [("项目A", [FakeStudent("1", "甲")]), ("项目B", [FakeStudent("2", "乙")])],
            strictness="high",
            comment_picker=scoring.CommentPicker(),
        )
        self.assertEqual(stats["graded"], 2)
        self.assertEqual(len(api.audit_calls), 2)

    def test_grade_groups_default_stats_dict(self):
        stats = GradingService(FakeApiClient()).grade_groups(
            [("项目A", [FakeStudent("1", "张三")])],
            strictness="high",
            comment_picker=scoring.CommentPicker(),
        )
        for key in ("graded", "failed", "min_score_names", "failed_names"):
            self.assertIn(key, stats)

    def test_grade_groups_respects_passed_in_stats(self):
        api = FakeApiClient()
        pre = {"total": 0, "graded": 5, "failed": 0, "min_score_names": [], "failed_names": []}
        stats = GradingService(api).grade_groups(
            [("项目A", [FakeStudent("1", "张三")])],
            strictness="high",
            comment_picker=scoring.CommentPicker(),
            stats=pre,
        )
        self.assertEqual(stats["graded"], 6)  # 5 + 1


class TemplateServiceTests(_TemplatesIsolatedTest):
    def test_list_all_returns_pools(self):
        templates = TemplateService().list_all()
        self.assertIn("short", templates)
        self.assertIn("long", templates)

    def test_add_appends_to_pool(self):
        svc = TemplateService()
        svc.add("short", "新增短评")
        templates = svc.list_all()
        self.assertIn("新增短评", templates["short"])
        self.assertEqual(len(templates["short"]), 3)

    def test_edit_replaces_at_index(self):
        svc = TemplateService()
        svc.edit("short", 0, "改后内容")
        self.assertEqual(svc.get("short", 0), "改后内容")
        self.assertEqual(svc.get("short", 1), "默认短评2")  # 其他不变

    def test_delete_removes_at_index(self):
        svc = TemplateService()
        ok = svc.delete("short", 0)
        self.assertTrue(ok)
        self.assertEqual(len(svc.list_all()["short"]), 1)

    def test_delete_refused_when_only_one_left(self):
        svc = TemplateService()
        ok = svc.delete("long", 0)  # long 池只有 1 条
        self.assertFalse(ok)
        self.assertEqual(len(svc.list_all()["long"]), 1)

    def test_get_returns_none_for_out_of_range(self):
        self.assertIsNone(TemplateService().get("short", 99))


class ExcelExporterTests(unittest.TestCase):
    def setUp(self):
        fd, path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        self._tmp_path = path

    def tearDown(self):
        try:
            os.remove(self._tmp_path)
        except OSError:
            pass

    def test_export_writes_rows_sorted_desc(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl 未安装")
        rows = [
            SimpleNamespace(student_name="张三", pro_score=80),
            SimpleNamespace(student_name="李四", pro_score=95),
            SimpleNamespace(student_name="王五", pro_score=70),
        ]
        ExcelExporter.export(rows, self._tmp_path)
        wb = load_workbook(self._tmp_path)
        ws = wb.active
        self.assertEqual(ws.max_row, 4)  # 表头 + 3 行
        self.assertEqual(ws.cell(1, 1).value, "姓名")
        self.assertEqual(ws.cell(1, 2).value, "分数")
        # 降序：李四(95) 第一行数据
        self.assertEqual(ws.cell(2, 1).value, "李四")
        self.assertEqual(ws.cell(2, 2).value, 95)
        self.assertEqual(ws.cell(4, 1).value, "王五")  # 70 最后


if __name__ == "__main__":
    unittest.main()
